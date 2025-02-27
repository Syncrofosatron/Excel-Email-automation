from email import generator
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl as excel
from openpyxl import Workbook
from openpyxl.styles import Border, Side

# This variable is useful, it will be equal to the number of distinct Clients we have.
client_number = 0

# workbook is an Excel file.
# We are loading this file using the name and extension.
workbook = excel.load_workbook("TestFile.xlsx")

# worksheet are the various sheets we have within workbook, default is 1.
# From the workbook that we have loaded, we are using the first worksheet.
worksheet = workbook.worksheets[0]

# cellA to cellJ are the respective columns, storing them on variables.
cellA = worksheet['A']
cellB = worksheet['B']
cellC = worksheet['C']
cellE = worksheet['E']
cellF = worksheet['F']
cellI = worksheet['I']
cellJ = worksheet['J']

# Take all the individual Client names (non-redundant) from column C and put in a list as declared.
client_name = []

for cell in cellC:
    if cell.value not in client_name:
        client_name.append(cell.value)

# Sort the client names alphabetically.
client_name.sort()

# Making more list variables to store the data of respective columns.
cellA_store = []
cellB_store = []
cellC_store = []
cellE_store = []
cellF_store = []
cellI_store = []
cellJ_store = []

# Since the first row was not included, I manually did here.
cellA_store.append(cellA[0].value)
cellB_store.append(cellB[0].value)
cellC_store.append(cellC[0].value)
cellI_store.append(cellI[0].value)
cellJ_store.append(cellJ[0].value)

# This will be equal to the length of the number of clients, and will increment from 0 to number of Clients.
counter = 0

while counter < len(client_name):

    for i in range(len(cellC)):
        if cellC[i].value and cellC[i].value == client_name[counter]:
            cellA_store.append(cellA[i].value)
            cellB_store.append(cellB[i].value)
            cellC_store.append(cellC[i].value)
            cellI_store.append(cellI[i].value)
            cellJ_store.append(cellJ[i].value)

# Changes the none in cells to underscore.

    for i in range(len(cellI_store)):
        if cellI_store[i] is None:
            cellI_store[i] = '_'

    for i in range(len(cellJ_store)):
        if cellJ_store[i] is None:
            cellJ_store[i] = '_'

# This code is responsible to create a mail draft.
# Give Subject, To and Body field information below.
# Note: Keep the Excel file to import from in the same directory as this  python script.

    new_workbook = excel.load_workbook("Employee Manager.xlsx")
    current_worksheet = new_workbook.active

# This is where we will get all the content we extracted and put in an email draft file.
    def create_email():
        msg = MIMEMultipart("alternative")
        msg["Subject"] = client_name[counter] + ": B&QRs."
        column_value = client_name[counter]
        values = []
        for n_row in current_worksheet.iter_rows():
            if n_row[0].value == column_value:
                values.append(n_row[1].value)

        # This will remove redundant mail addresses from values.
        values = [x for k, x in enumerate(values) if x not in values[:k]]

        for k in range(len(values)):
            msg["To"] = values[k]
        msg["Cc"] = "abc@example.com, bcd@example.com"

        html = f"""\
        <html>
        <head></head>
        <style>
        td {{
        border: 1px solid black;
        }}
        table {{
        border-collapse: collapse
        }}
        </style>
        <body>Hello,
    
    Please consider it on urgent basis.<br><br>
    We have the following details and we want you to ensure:<br><br>
    <table>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellA_store)}</table>
    </td>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellB_store)}</table>
    </td>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellC_store)}</table>
    </td>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellI_store)}</table>
    </td>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellJ_store)}</table>
    </td>
    </table>
<br>    Regards
<br>Neeraj Mishra
<br>
    </body>
        </html>"""

        part = MIMEText(html, "html")
        msg.attach(part)
        msg.add_header("X-Unsent", '1')

    # Remember to remove "/" symbol, or any "special characters" from Client name or the script will stop.
        outfile_name = r"C:/Users/generated mails/" + client_name[counter] + ".eml"
        with open(outfile_name, 'w') as outfile:
            gen = generator.Generator(outfile)
            gen.flatten(msg)

# Calling the function above to generate all the mail draft files.
    create_email()

# Emptying the data so as to fill new data and avoid redundancy.
    cellA_store = []
    cellB_store = []
    cellC_store = []
    cellI_store = []
    cellJ_store = []

# Adding manually the header.
    cellA_store.append(cellA[0].value)
    cellB_store.append(cellB[0].value)
    cellC_store.append(cellC[0].value)
    cellI_store.append(cellI[0].value)
    cellJ_store.append(cellJ[0].value)

    counter += 1

    # POSSIBLE IMPROVEMENTS:
    # 0. GUI (Graphical User Interface).
    # 1. Separating the data based on the contact. - Maybe store the contact in separate
    #    file and get the emails from there?
    # 2. To reduce the step of manually creating the "TestFile.xlsx".