from email import generator
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import numpy as np

import openpyxl as excel
from openpyxl import Workbook
from openpyxl.styles import Border, Side

# This variable is useful, it will be equal to the number of distinct Clients we have.
client_number = 0

# workbook is an Excel file.
# We are loading this file using the name and extension.
workbook = excel.load_workbook("TestFile.xlsx")

# worksheet are the various sheets we have within workbook, default is 1.
# From the workbook that we have loaded, we are using the first worksheet.
worksheet = workbook.worksheets[0]

# cellA to cellBE are the respective columns, storing them on variables.
cellA = worksheet['A']
cellB = worksheet['B']
cellC = worksheet['C']
cellD = worksheet['D']
cellE = worksheet['E']
cellF = worksheet['F']
cellG = worksheet['G']
cellH = worksheet['H']
cellJ = worksheet['J']
cellBE = worksheet['BE']

# Take all the individual Client names (non-redundant) from column C and put in a list as declared.
client_name = []
manager_name = []
manager_mail = []

for cell in cellH:
    if cell.value not in client_name:
        client_name.append(cell.value)

for cell in cellBE:
    if cell.value not in manager_mail:
        manager_mail.append(cell.value)

# Since the last element in the list is None or empty cell, just popping it out.
# Note: If any client name or mail seems to be missing, probably check here.
client_name.pop()
manager_mail.pop()

# Sort the client names alphabetically and managers mail.
client_name.sort()
manager_mail.sort()

# Making more list variables to store the data of respective columns.
cellA_store = []
cellB_store = []
cellC_store = []
cellD_store = []
cellE_store = []
cellF_store = []
cellG_store = []
cellH_store = []
cellJ_store = []
cellBE_store = []

# Since the first row was not included, I manually did here.
cellA_store.append(cellA[0].value)
cellB_store.append(cellB[0].value)
cellC_store.append(cellC[0].value)
cellD_store.append(cellD[0].value)
cellE_store.append(cellE[0].value)
cellF_store.append(cellF[0].value)
cellG_store.append(cellG[0].value)
cellH_store.append(cellH[0].value)
cellJ_store.append(cellJ[0].value)
#cellBE_store.append(cellBE[0].value)

# This will be equal to the length of the number of clients, and will increment from 0 to number of Clients.
counter = 0

while counter < len(client_name):

    for i in range(len(cellH)):
        if cellH[i].value and cellH[i].value == client_name[counter]:
            cellA_store.append(cellA[i].value)
            cellB_store.append(cellB[i].value)
            cellC_store.append(cellC[i].value)
            cellD_store.append(cellD[i].value)
            cellE_store.append(cellE[i].value)
            cellF_store.append(cellF[i].value)
            cellG_store.append(cellG[i].value)
            cellH_store.append(cellH[i].value)
            cellJ_store.append(cellJ[i].value)

    for i in range(len(cellBE)):
        if cellBE[i].value and cellBE[i].value == manager_mail[counter]:
            cellBE_store.append(cellBE[i].value)

    # This changes the "None" field in empty cells to underscore.
    for i in range(len(cellJ_store)):
        if cellJ_store[i] is None:
            cellJ_store[i] = "..."

    # Creating new workbook to just have column H and column BE.
    new_workbook = Workbook()
    current_worksheet = new_workbook.active

    for row in range(4, 34):
        new_workbook_cellH = worksheet.cell(row=row,column=8)
        new_workbook_cellBE = worksheet.cell(row=row, column=57)
        current_worksheet.cell(row=row-3, column=1, value=new_workbook_cellH.value)
        current_worksheet.cell(row=row-3, column=2, value=new_workbook_cellBE.value)

    new_workbook.save("Client Manager EMails.xlsx")

    def create_email():
        msg = MIMEMultipart("alternative")
        msg["Subject"] = client_name[counter] + ": Expiring IDs."

        # Holy shit this is working!
        column_value = client_name[counter]
        values = []
        for n_row in current_worksheet.iter_rows():
            if n_row[0].value == column_value:
                values.append(n_row[1].value)

        # This will remove redundant mail addresses from values.
        values = [x for k, x in enumerate(values) if x not in values[:k]]

        for k in range(len(values)):
            msg["To"] = values[k]
        msg["Cc"] = "abc@example.com, bcd@example.com"

        html = f"""\
        <html>
        <head></head>
        <style>
        td {{
        border: 1px solid black;
        }}
        table {{
        border-collapse: collapse
        }}
        </style>
        <body>Hello,<br><br>
The following position(s) are expiring this month, could you please verify and confirm?
<br><br>
    <table>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellA_store)}</table>
    </td>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellB_store)}</table>
    </td>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellC_store)}</table>
    </td>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellD_store)}</table>
    </td>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellE_store)}</table>
    </td>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellF_store)}</table>
    </td>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellG_store)}</table>
    </td>
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellH_store)}</table>
    </td>                
    <td>
    <table>{''.join(f'<tr><td>{item}</td></tr>' for item in cellJ_store)}</table>
    </td>
    </table>
<br>    Regards
<br>Neeraj Mishra
<br>
    </body>
        </html>"""

        part = MIMEText(html, "html")
        msg.attach(part)
        msg.add_header("X-Unsent", '1')

        outfile_name = r"C:/Users/generated mails/" + client_name[
            counter] + ".eml"
        with open(outfile_name, 'w') as outfile:
            gen = generator.Generator(outfile)
            gen.flatten(msg)


    # Calling the function above to generate all the mail draft files.
    create_email()

    # Emptying the data so as to fill new data and avoid redundancy.
    cellA_store = []
    cellB_store = []
    cellC_store = []
    cellD_store = []
    cellE_store = []
    cellF_store = []
    cellG_store = []
    cellH_store = []
    cellJ_store = []

    # Adding manually the header.
    cellA_store.append(cellA[0].value)
    cellB_store.append(cellB[0].value)
    cellC_store.append(cellC[0].value)
    cellD_store.append(cellD[0].value)
    cellE_store.append(cellE[0].value)
    cellF_store.append(cellF[0].value)
    cellG_store.append(cellG[0].value)
    cellH_store.append(cellH[0].value)
    cellJ_store.append(cellJ[0].value)

    counter += 1